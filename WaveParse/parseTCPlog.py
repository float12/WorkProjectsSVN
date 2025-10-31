# -*- coding: gbk -*-
import re
from datetime import datetime
from collections import deque
import pandas as pd
import matplotlib.pyplot as plt
import openpyxl
import sys
from PyQt5.QtWidgets import QComboBox, QHBoxLayout, QWidget
from PyQt5.QtWidgets import (
    QApplication, QWidget, QVBoxLayout, QPushButton, QFileDialog, QLabel, QMessageBox
)
from PyQt5.QtCore import Qt

def diff_ms(dt1: datetime, dt2: datetime) -> int:
    """
    �������� datetime �ĺ���� (dt2 - dt1)
    ������������
    """
    delta = dt2 - dt1
    return int(delta.total_seconds() * 1000)

class LogAnalyzer:
    def __init__(self, log_path):
        self.log_path = log_path
        self.entries = []
        self.matched = []
        self.no_reply = []
        self.unmatched_recvs = []
        self.df = pd.DataFrame()
        self.send_flag = 0
        self.timeGap = []
        self.No_reply_time = []
        self.send_time = None

    def parse_log(self, time_format="fmt1"):
        """
        ������־��֧�����ָ�ʽ��ͨ��time_format��������:
        fmt1: [15:53:42.790]������.../�ա���...
        fmt2: TX:.../RX:...
        fmt3: [2025-09-01 16:08:03.252]# SEND HEX.../RECV HEX...
        """
        self.timeGap.clear()
        self.No_reply_time.clear()
        self.send_flag = 0
        self.send_time = None

        if time_format == "fmt1":
            # ��һ�ָ�ʽ
            time_re = re.compile(r'\[(\d{2}:\d{2}:\d{2}\.\d{3})\]')
            with open(self.log_path, 'r', encoding='gbk') as f:
                for i, line in enumerate(f, 1):
                    m = time_re.search(line)
                    if not m:
                        continue
                    timestr = m.group(1)
                    today = datetime.now().date()
                    dt = datetime.strptime(f"{timestr}", '%H:%M:%S.%f')
                    tail = line[m.end():]
                    if '��' in tail and self.send_flag == 0:
                        self.send_flag = 1
                        self.send_time = dt
                    elif '��' in tail and self.send_flag == 1:
                        self.send_flag = 0
                        self.timeGap.append(diff_ms(self.send_time, dt))
                    elif '��' in tail and self.send_flag == 1:
                        print(f"send without response, time {self.send_time}")
                        self.send_time = dt
                        self.No_reply_time.append(self.send_time)
                        continue
                    else:
                        continue

        elif time_format == "fmt2":
            # �ڶ��ָ�ʽ
            # ����17:21:37:419 RX:FE FE FE FE ...
            time_re = re.compile(r'(\d{2}:\d{2}:\d{2}:\d{3})\s+(TX|RX):')
            with open(self.log_path, 'r', encoding='gbk') as f:
                for i, line in enumerate(f, 1):
                    m = time_re.search(line)
                    if not m:
                        continue
                    timestr, direction = m.group(1), m.group(2)
                    today = datetime.now().date()
                    # 17:21:37:419 �� %H:%M:%S:%f
                    dt = datetime.strptime(f"{today} {timestr}", '%Y-%m-%d %H:%M:%S:%f')
                    if direction == "TX" and self.send_flag == 0:
                        self.send_flag = 1
                        self.send_time = dt
                    elif direction == "RX" and self.send_flag == 1:
                        self.send_flag = 0
                        self.timeGap.append(diff_ms(self.send_time, dt))
                    elif direction == "TX" and self.send_flag == 1:
                        print(f"send without response, time {self.send_time}")
                        self.send_time = dt
                        self.No_reply_time.append(dt)
                        continue
                    else:
                        continue

        elif time_format == "fmt3":
            # �����ָ�ʽ
            # [2025-09-01 16:08:03.252]# SEND HEX/12 >>>
            # [2025-09-01 16:08:03.314]# RECV HEX/109 from SERVER <<<
            time_re = re.compile(r'\[(\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}\.\d{3})\].*# (SEND|RECV) HEX')
            with open(self.log_path, 'r', encoding='gbk') as f:
                for i, line in enumerate(f, 1):
                    m = time_re.search(line)
                    if not m:
                        continue
                    timestr, direction = m.group(1), m.group(2)
                    dt = datetime.strptime(timestr, '%Y-%m-%d %H:%M:%S.%f')
                    if direction == "SEND" and self.send_flag == 0:
                        self.send_flag = 1
                        self.send_time = dt
                    elif direction == "RECV" and self.send_flag == 1:
                        self.send_flag = 0
                        self.timeGap.append(diff_ms(self.send_time, dt))
                    elif direction == "SEND" and self.send_flag == 1:
                        print(f"send without response, time {self.send_time}")
                        self.send_time = dt
                        self.No_reply_time.append(dt)
                        continue
                    else:
                        continue

        else:
            print("��֧�ֵ���־��ʽ��������ѡ��fmt1/fmt2/fmt3")

    def export_excel(self, path: str):
        """�� timeGap �б����� Excel��ÿһ��һ��һ�����ݣ����� No_reply_time ��������һ��sheet"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Intervals"
        ws.cell(row=1, column=1, value="interval_ms")  # ��ͷ

        for i, val in enumerate(self.timeGap, start=2):
            ws.cell(row=i, column=1, value=val)

        # �½�һ��sheet�����޻ظ�ʱ���
        ws2 = wb.create_sheet(title="�޻ظ�ʱ���")
        ws2.cell(row=1, column=1, value="no_reply_time")  # ��ͷ
        for i, dt in enumerate(self.No_reply_time, start=2):
            # ��ʽ��Ϊ�ַ���д��
            ws2.cell(row=i, column=1, value=str(dt))

        wb.save(path)
        print(f"Excel �ѱ���: {path}")

    def plot_intervals(self, path):
        plt.figure(figsize=(8, 4))
        x = [i for i in range(len(self.timeGap))]
        plt.plot(x, self.timeGap, marker='o')
        plt.xlabel('send_time')
        plt.ylabel('interval (ms)')
        plt.title('Send��Recv intervals')
        plt.xticks(rotation=45)
        plt.tight_layout()
        plt.savefig(path)
        plt.show()

class MainWindow(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("TCP��־��������")
        self.setGeometry(300, 300, 400, 200)
        self.log_path = ""
        self.analyzer = None

        self.layout = QVBoxLayout()
        self.setLayout(self.layout)

        self.label = QLabel("��ѡ����־�ļ�")
        self.label.setAlignment(Qt.AlignCenter)
        self.layout.addWidget(self.label)

        self.btn_select = QPushButton("ѡ����־�ļ�")
        self.btn_select.clicked.connect(self.select_file)
        self.layout.addWidget(self.btn_select)

        # ����������ѡ��ʱ���ʽ
        self.format_group = QWidget()
        self.format_layout = QHBoxLayout()
        self.format_group.setLayout(self.format_layout)
        
        self.format_label = QLabel("��־��ʽ:")
        self.format_layout.addWidget(self.format_label)
        
        self.format_combo = QComboBox()
        self.format_combo.addItem("��ʽ1: [hh:mm:ss.xxx]������.../�ա���...", "fmt1")
        self.format_combo.addItem("��ʽ2: TX:.../RX:...", "fmt2")
        self.format_combo.addItem("��ʽ3: [yyyy-mm-dd hh:mm:ss.xxx]# SEND/RECV HEX...", "fmt3")
        self.format_layout.addWidget(self.format_combo)
        
        self.layout.addWidget(self.format_group)

        self.btn_parse = QPushButton("��ʼ����")
        self.btn_parse.clicked.connect(self.start_parse)
        self.btn_parse.setEnabled(False)
        self.layout.addWidget(self.btn_parse)
        self.result_label = QLabel("")
        self.result_label.setAlignment(Qt.AlignCenter)
        self.layout.addWidget(self.result_label)

    def select_file(self):
        file_path, _ = QFileDialog.getOpenFileName(self, "ѡ����־�ļ�", "", "��־�ļ� (*.log *.txt);;�����ļ� (*)")
        if file_path:
            self.log_path = file_path
            self.label.setText(f"��ѡ���ļ�: {file_path}")
            self.btn_parse.setEnabled(True)
            self.result_label.setText("")
        else:
            self.label.setText("δѡ���ļ�")
            self.btn_parse.setEnabled(False)

    def start_parse(self):
        if not self.log_path:
            QMessageBox.warning(self, "����", "����ѡ����־�ļ���")
            return
        try:
            self.analyzer = LogAnalyzer(self.log_path)
            # ��ȡѡ��ĸ�ʽ
            selected_format = self.format_combo.currentData()
            self.analyzer.parse_log(selected_format)
            # ����Excel
            excel_path = self.log_path + "_intervals.xlsx"
            self.analyzer.export_excel(excel_path)
            # ��ͼ
            img_path = self.log_path + "_intervals.png"
            self.analyzer.plot_intervals(img_path)
            self.result_label.setText(f"������ɣ�\nExcel: {excel_path}\nͼƬ: {img_path}")
        except Exception as e:
            QMessageBox.critical(self, "����ʧ��", f"��������: {e}")

if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec_())
