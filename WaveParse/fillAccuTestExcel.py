# -*- coding: gbk -*-
import re
from openpyxl import load_workbook, Workbook
from ParseUIData import FrameParser,WavePlotter
import os
from pathlib import Path
def parse_total_str(total_str):
    """解析原始数据字符串，返回各功率/电流/电压列表"""
    pattern = r"点\s+([-\d.\s]+?)(?=每|$)"
    matches = re.findall(pattern, total_str, re.S)
    data_lists = [list(map(float, m.split())) for m in matches]
    if len(data_lists) < 5:
        raise ValueError("数据解析失败，数据组数不足5")
    return {
        "appr_pwr_list": data_lists[0],
        "act_pwr_list": data_lists[1],
        "react_pwr_list": data_lists[2],
        "iRms_list": data_lists[3],
        "uRms_list": data_lists[4]
    }

def extract_same_data(data, difFactor, min_group_len=4, print_group=False,data_name =None):
    """分组提取相近数据，返回每组中间4个或全部（不足4个）"""
    groups = []
    current_group = [data[0]]
    for i in range(1, len(data)):
        if abs(data[i] - data[i-1]) < abs(data[i-1])*difFactor or abs(data[i] - data[i-1]) < 2:
            current_group.append(data[i])
        else:
            n = len(current_group)
            if n <= min_group_len:
                mid_nums = current_group[:]
            else:
                start = n // 2 - 2
                if start < 0: start = 0
                mid_nums = current_group[start:start + min_group_len]
            groups.append(mid_nums)
            current_group = [data[i]]

    # ? 处理最后一组
    n = len(current_group)
    if n <= min_group_len:
        mid_nums = current_group[:]
    else:
        start = n // 2 - 2
        if start < 0: start = 0
        mid_nums = current_group[start:start + min_group_len]
    groups.append(mid_nums)

    groups_filtered = [g for g in groups if len(g) >= min_group_len]
    if print_group:
        print(data_name)
        for g in groups_filtered:
            print(*g)
    return groups_filtered

def check_group_count(data_group, expect_count, name, writeExcelFlag):
    """检查分组数量是否符合预期"""
    if len(data_group) != expect_count:
        print(f"{name} len err")
        return 0
    return writeExcelFlag
#电动车录波写入excel
def write_to_excel_ddc(output_excel_name, output_sheetName, UIPoints, power_start_index,
                   appr_pwr_list, act_pwr_list, react_pwr_list):
    """写入数据到Excel"""
    wb = None
    try:
        try:
            wb = load_workbook(output_excel_name)
            sheet = wb[output_sheetName]
        except FileNotFoundError:
            wb = Workbook()
            sheet = wb.active
            sheet.title = output_sheetName

        for i in range(UIPoints):
            # 视在功率
            val = " ".join(f"{x:.2f}" for x in appr_pwr_list[power_start_index + i])
            sheet.cell(row=2 + i, column=3, value=val)
            # 有功
            val = " ".join(f"{x:.2f}" for x in act_pwr_list[power_start_index + i])
            sheet.cell(row=2 + i, column=7, value=val)
            # 无功
            val = " ".join(f"{x:.2f}" for x in react_pwr_list[power_start_index + i])
            sheet.cell(row=2 + i, column=11, value=val)
    finally:
        if wb:
            wb.save(output_excel_name)
            wb.close()
def write_to_excel_waveRecord(output_excel_name, output_sheetName, UIPoints, power_start_index,
                   appr_pwr_list, act_pwr_list, react_pwr_list, uRms_list, iRms_list):
    """写入数据到Excel"""
    wb = None
    try:
        try:
            wb = load_workbook(output_excel_name)
            sheet = wb[output_sheetName]
        except FileNotFoundError:
            wb = Workbook()
            sheet = wb.active
            sheet.title = output_sheetName

        for i in range(2, 2+UIPoints):
            # if i != 2:
                # U RMS
                # val = " ".join(map(str, uRms_list[i - 3]))
                # sheet.cell(row=i+8, column=2, value=val)
                # I RMS
                # val = " ".join(map(str, iRms_list[i - 3]))
                # sheet.cell(row=i+8, column=6, value=val)

            # 视在功率
            val = " ".join(f"{x:.2f}" for x in appr_pwr_list[power_start_index + i - 2])
            sheet.cell(row=9 - i, column=5, value=val)
            # 1.0有功
            val = " ".join(f"{x:.2f}" for x in act_pwr_list[power_start_index + (i - 2)*3])
            sheet.cell(row=9 - i, column=5 + 6, value=val)
            # 1.0无功
            val = " ".join(f"{x:.2f}" for x in react_pwr_list[power_start_index + (i - 2)*3])
            sheet.cell(row=9 - i, column=5 + 6 + 5, value=val)
            # 0.5有功
            val = " ".join(f"{x:.2f}" for x in act_pwr_list[power_start_index + (i - 2)*3 + 1])
            sheet.cell(row=9 - i, column=5 + 6 + 5 + 5, value=val)
            # 0.5无功
            val = " ".join(f"{x:.2f}" for x in react_pwr_list[power_start_index + (i - 2)*3 + 1])
            sheet.cell(row=9 - i, column=5 + 6 + 5 + 5 + 5, value=val)
            # 0.8有功
            val = " ".join(f"{x:.2f}" for x in act_pwr_list[power_start_index + (i - 2)*3 + 2])
            sheet.cell(row=9 - i, column=5 + 6 + 5 + 5 + 5 + 5, value=val)
            # 0.8无功
            val = " ".join(f"{x:.2f}" for x in react_pwr_list[power_start_index + (i - 2)*3 + 2])
            sheet.cell(row=9 - i, column=5 + 6 + 5 + 5 + 5 + 5 + 5, value=val)
    finally:
        if wb:
            wb.save(output_excel_name)
            wb.close()
def get_uifactor(uifactor_sheet_path,type):
    """
    读取指定sheet的F列芯片相关名称，返回为列表（从F2开始，不含表头）
    """
    factors = []
    wb = load_workbook(uifactor_sheet_path, read_only=True)
    if type =="ddc":
        sheet = wb["数据块"]
        # F列为第6列
        row = 2
        while True:
            if (row - 1) % 7 == 0:
                row += 1
                continue
            value = sheet.cell(row=row, column=6).value
            if value is None:
                break
            factors.append(float(value))
            row += 1
        wb.close()
    elif type == "waveRecord":#物联表录波格式
        None
    return factors


def natural_key(s):
    parts = re.split(r'(\d+)', s)
    key = []
    for part in parts:
        if part.isdigit():
            key.append(int(part))
        else:
            key.append(part.lower())
    return key
def average_every_50(nums):
    """对列表每100个数字取平均，保存到源列表。原地修改。"""
    avg_list = []
    n = len(nums)
    for i in range(0, n, 50):
        chunk = nums[i:i+50]
        if len(chunk) > 0:
            avg = sum(chunk) / len(chunk)
            avg_list.append(avg)
    return avg_list
def process_and_write(parser, output_excel_name, output_sheetName, UIPoints, power_start_index,type):
    # 解析数据
    appr_pwr_avg_list = average_every_50(parser.plot_AppPwr_data)
    act_pwr_avg_list = average_every_50(parser.plot_ActPwr_data)
    react_pwr_avg_list = average_every_50(parser.plot_ReactPwr_data)
    uRms_avg_list = average_every_50(parser.plot_UA_rms_data)
    iRms_avg_list = average_every_50(parser.plot_IA_rms_data)

    appr_pwr_list = extract_same_data(appr_pwr_avg_list, 0.04, print_group=True,data_name = "视在")
    act_pwr_list = extract_same_data(act_pwr_avg_list, 0.04, print_group=True,data_name = "有功")
    react_pwr_list = extract_same_data(react_pwr_avg_list, 0.06, print_group=True,data_name = "无功")
    appr_pwr_list = appr_pwr_list[:4] if len(appr_pwr_list) >= 5 else appr_pwr_list
    act_pwr_list = act_pwr_list[:4] if len(act_pwr_list) >= 5 else act_pwr_list
    react_pwr_list = react_pwr_list[:4] if len(react_pwr_list) >= 5 else react_pwr_list
    if type == "waveRecord":
        uRms_list = extract_same_data(uRms_avg_list, 0.05, print_group=False,data_name ="urms")
        iRms_list = extract_same_data(iRms_avg_list, 0.05, print_group=False,data_name ="Irms")
        writeExcelFlag = 1
        writeExcelFlag = check_group_count(appr_pwr_list, power_start_index+UIPoints, "appr_pwr_list", writeExcelFlag)
        writeExcelFlag = check_group_count(act_pwr_list, power_start_index+UIPoints*3, "act_pwr_list", writeExcelFlag)
        writeExcelFlag = check_group_count(react_pwr_list, power_start_index+UIPoints*3, "react_pwr_list", writeExcelFlag)
        if writeExcelFlag:
            write_to_excel_waveRecord(
                output_excel_name, output_sheetName, UIPoints, power_start_index,
                appr_pwr_list, act_pwr_list, react_pwr_list, uRms_list, iRms_list
            )
    elif type == "ddc":
        writeExcelFlag = 1
        writeExcelFlag = check_group_count(appr_pwr_list, power_start_index + UIPoints, "appr_pwr_list", writeExcelFlag)
        writeExcelFlag = check_group_count(act_pwr_list, power_start_index + UIPoints , "act_pwr_list",
                                           writeExcelFlag)
        writeExcelFlag = check_group_count(react_pwr_list, power_start_index + UIPoints , "react_pwr_list",
                                           writeExcelFlag)
        if writeExcelFlag:
            write_to_excel_ddc(
                output_excel_name, output_sheetName, UIPoints, power_start_index,
                appr_pwr_list, act_pwr_list, react_pwr_list
            )
    return writeExcelFlag
# ----------------- 主流程入口 -----------------
if __name__ == "__main__":
    output_sheetName = "1"
    output_excel_name = r"D:\work\project\ElectricityCycle\电动车精度数据\电动车录波精度点检表 - 副本 (6).xlsx"
    data_file_dir_path = r"D:\work\project\ElectricityCycle\电动车精度数据\6"
    uifactor_sheet_path = r"D:\work\project\ElectricityCycle\0006系数.xlsx"
    power_start_index = 0#功率开始索引
    UIPoints = 4#UI检定点数，ddc：每个检表点，waveRecord：1.0 0.5l 0.8c算一个点
    type = "ddc" #ddc waveRecord
    loop_num = 10


    #读取uifactor
    oneLoop_uifactor = [0, 0, 0, 0, 0, 0]
    allLoop_uifactor = get_uifactor(uifactor_sheet_path,type)
    AllLoop_data_file_path = []
    #读取数据文件
    dir_path = Path(data_file_dir_path)
    all_files = sorted([p for p in dir_path.iterdir() if p.is_file()], key=lambda p: natural_key(p.name))
    AllLoop_data_file_path.extend(all_files)
    
    #0 有序号的物联表录波格式 1 698格式 2:物联表录波格式

    for i in range(loop_num):
        oneLoop_uifactor = list(oneLoop_uifactor)
        oneLoop_uifactor[0] = allLoop_uifactor[i // 3*6 + i%3]
        oneLoop_uifactor[3] = allLoop_uifactor[i // 3*6 + i%3 +3]
        # print(oneLoop_uifactor)
        # 获取目录中所有文件，并根据文件名排序
        oneLoop_uifactor = tuple(oneLoop_uifactor)
        parser = WavePlotter(phase=1, type=1,FilePath=AllLoop_data_file_path[i],uiFactor=oneLoop_uifactor,endian="big")
        parser.parse_and_fill_plot_data()

        if type == "ddc":
            output_sheetName = str(i + 1)
        elif type == "waveRecord":
            None
        if process_and_write(parser,output_excel_name,output_sheetName,UIPoints,power_start_index,type) == 1:
            print("write excel loop",i+1,"done")
        else:
            print("write excel loop",i+1,"error")