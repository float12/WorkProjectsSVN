# -*- coding: gbk -*-
import openpyxl
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
import os
import pandas as pd
import re
def extract_number(filename):
    nums = re.findall(r'\((\d+)\)', filename)
    return int(nums[-1]) if nums else 0


def merge_sheets_by_name(input_dir, output_file="merged.xlsx"):
    # 获取目录下所有 Excel
    files = [
        f for f in os.listdir(input_dir)
        if f.lower().endswith((".xlsx", ".xls"))
    ]
    files.sort(key=extract_number)

    # 读取第一个文件用于获取 sheet 名列表
    first_file = os.path.join(input_dir, files[0])
    sheet_names = pd.ExcelFile(first_file).sheet_names

    # 一个 dict：每个 sheet 一个列表，用于存放不同文件的 df
    merged = {sheet: [] for sheet in sheet_names}

    for file in files:
        file_path = os.path.join(input_dir, file)
        print(f"Processing: {file}")

        xls = pd.ExcelFile(file_path)

        for sheet in sheet_names:
            df = xls.parse(sheet)

            # 在开头插入文件名 + 空行
            head = pd.DataFrame([[file]])
            empty = pd.DataFrame([[""]])

            merged[sheet].append(head)
            merged[sheet].append(empty)
            merged[sheet].append(df)
            merged[sheet].append(empty)  # 文件区块结束再加空行

    # 写入输出文件
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        for sheet in sheet_names:
            result = pd.concat(merged[sheet], ignore_index=True)
            result.to_excel(writer, sheet_name=sheet, index=False)

    print(f"合并完成！输出文件：{output_file}")
def fill_wave_err_to_overview(excel_path):
    # 打开excel文件
    wb = openpyxl.load_workbook(excel_path, data_only=True)  # data_only=True 只读取值，不读取公式
    # 获取所有以"表"开头的sheet名
    table_sheets = [name for name in wb.sheetnames if name.startswith("表")]
    # 获取总览sheet
    overview_sheet = wb["录波精度总览"]

    # 需要读取的区域及其目标区域
    src_dst_map = [
        (("H2", "H7"), ("C3", "C8")),
        (("N2", "N7"), ("C10", "C15")),
        (("S2", "S7"), ("C17", "C22")),
        (("X2", "X7"), ("C24", "C29")),
        (("AC2", "AC7"), ("C31", "C36")),
        (("AH2", "AH7"), ("C38", "C43")),
        (("AM2", "AM7"), ("C45", "C50")),
        (("D10", "D15"), ("C53", "C58")),
        (("H10", "H15"), ("C61", "C66")),
    ]

    # 计算每个区域的行数
    def get_row_range(cell1, cell2):
        col1, row1 = coordinate_from_string(cell1)
        col2, row2 = coordinate_from_string(cell2)
        return int(row1), int(row2)

    # 每个表写入的起始列（C列，3号列）
    start_col = 3

    for idx, sheet_name in enumerate(table_sheets):
        sheet = wb[sheet_name]
        col_offset = idx  # 每个表右移一列

        for (src_range, dst_range) in src_dst_map:
            src_col, src_row_start = coordinate_from_string(src_range[0])
            _, src_row_end = coordinate_from_string(src_range[1])
            dst_col, dst_row_start = coordinate_from_string(dst_range[0])
            _, dst_row_end = coordinate_from_string(dst_range[1])

            # 计算目标列
            dst_col_num = column_index_from_string(dst_col) + col_offset
            # 读取源数据
            for i, row in enumerate(range(int(src_row_start), int(src_row_end)+1)):
                value = sheet[f"{src_col}{row}"].value
                # 写入目标sheet
                overview_sheet.cell(row=int(dst_row_start)+i, column=dst_col_num, value=value)

    wb.save(excel_path)
    print("数据已成功填入'录波精度总览'子表。")

if __name__ == "__main__":
    # fill_wave_err_to_overview("C:\\Users\\nh\\Desktop\\三相录波工装点检表250816 .xlsx")
    directory = r"D:\work\project\ElectricityCycle\电动车精度数据\tmp"
    merge_sheets_by_name(directory)
